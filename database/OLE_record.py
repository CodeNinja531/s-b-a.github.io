import sqlite3
import pandas as pd
import collections
import os
import openpyxl
import re

# Database path
DB_PATH = "database/Sports day helper"

# Connect
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

# SQL query
query = """
SELECT
    stu_info.class AS Class,
    stu_info.clno AS Clno,
    stu_info.name AS Name,
    event.item AS Event,
    event.gender AS Gender,
    event.grade AS Grade,
    leaderboard.rank AS Ranking
FROM leaderboard
JOIN participants ON leaderboard.athlete_id = participants.athlete_id
 AND leaderboard.event_id = participants.event_id
JOIN stu_info ON participants.stu_id = stu_info.stu_id
JOIN event ON leaderboard.event_id = event.event_id
ORDER BY stu_info.class, stu_info.clno, stu_info.name, leaderboard.rank
"""

cursor.execute(query)
rows = cursor.fetchall()

# Helper to convert rank to award string
def rank_to_award(rank):
    if rank == 1:
        return "Champion"
    elif rank == 2:
        return "1st Runner-up"
    elif rank == 3:
        return "2nd Runner-up"
    elif rank == 4:
        return "4th Place"
    elif rank == 5:
        return "5th Place"
    else:
        return f"{rank}th Place"


def class_sort_key(class_name):
    # Assumes class format like '1A', '2B', etc.
    # Splits into (grade:int, section:str)
    m = re.match(r"(\d+)([A-D])", class_name)
    grade = int(m.group(1))
    section = m.group(2)
    return (grade, section)


student_awards = collections.defaultdict(list)
for row in rows:
    key = (row[0], row[1], row[2])  # (Class_name, clno, name)
    gender = row[4]
    if gender.lower() == "boy":
        gender_str = "Boy's"
    elif gender.lower() == "girl":
        gender_str = "Girl's"
    else:
        gender_str = f"{gender}'s"
    event_full = f"{gender_str} {row[5]} Grade {row[3]}"
    award_str = rank_to_award(row[6])
    student_awards[key].append((award_str, event_full))


# list of data
excel_rows = []
sorted_students = sorted(student_awards.items(), key=lambda x: (class_sort_key(x[0][0]), int(x[0][1]), x[0][2]))
for (class_name, clno, name), awards in sorted_students:
    for award, event_name in awards:
        excel_rows.append([class_name, clno, name, 'participant', f"{award} in {event_name}"])
    excel_rows.append(['', '', '', '', ''])

# Output file name
output_file = 'OLE.xlsx'
if os.path.exists(output_file):
    os.remove(output_file)

# Data input to Excel
df = pd.DataFrame(excel_rows, columns=['Class', 'Clno', 'Name', 'Role', 'Award'])

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    pd.DataFrame([["Sports Day OLE Record"]]).to_excel(writer, index=False, header=False)
    df.to_excel(writer, index=False, startrow=1)

# Data formatting in Excel
wb = openpyxl.load_workbook(output_file)
ws = wb.active
ws.merge_cells('A1:E1')

ws['A1'].font = openpyxl.styles.Font(size=20, bold=True)

ws.freeze_panes = 'A3'

for col in range(1, 6):
    max_length = 0
    col_letter = openpyxl.utils.get_column_letter(col)
    for cell in ws[col_letter]:
        try:
            cell_value = str(cell.value)
        except:
            cell_value = ""
        if cell_value:
            max_length = max(max_length, len(cell_value))
    ws.column_dimensions[col_letter].width = max_length + 2

# Save and leave
wb.save(output_file)
conn.close()